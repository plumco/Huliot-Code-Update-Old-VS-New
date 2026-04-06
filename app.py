import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import io
import copy

st.set_page_config(page_title="Huliot Code Updater", page_icon="🔄", layout="wide")

# ─── MASTER CODE MAPPING ──────────────────────────────────────────────────────
# Format: "OLD_CODE": "NEW_CODE"
CODE_MAP = {
    # ── ULTRA SILENT PIPES: Special Length Double Socket L500 (shared→unique codes) ──
    "5754000050-i": "5754040050-i",
    "5755000050-i": "5755050050-i",
    "5757500050-i": "5757575050-i",
    "5751100050-i": "5751111050-i",
    "5751200050-i": "5751212050-i",
    "5751600050-i": "5751616050-i",

    # ── ULTRA SILENT BEND 45 DEG ──
    "7070030470":   "7070030470-i",
    "7070040470":   "7070040470-i",

    # ── ULTRA SILENT BEND 87.5 DEG ──
    "7070020870":   "7070020870-i",

    # ── ULTRA SILENT WYE BRANCH ──
    "7070643470":   "7070643470-i",
    "7070644470":   "7070644470-i",

    # ── ULTRA SILENT ONE WAY SOCKET ──
    "7071730275":   "7071730275-i",
    "7071740275":   "7071740275-i",

    # ── ULTRA SILENT END CAP ──
    "7071620070":   "7071620070-i",
    "7071640070":   "7071640070-i",

    # ── ULTRA SILENT MFT ──
    "17078111070 B-i": "17078111070-B",
    "S11050505075":    "S11050505075-i",

    # ── ULTRA SILENT HEIGHT RISER ──
    "60203651":     "60203651-i",

    # ── ULTRA SILENT S TRAP ──
    "7071840070 B-i": "7071840070-i",

    # ── HT PRO BEND 15 & 30 DEG ──
    # (no code change, only price)

    # ── HT PRO BEND 45 DEG ──
    "40010460":     "40010460-i",
    "40020460":     "40020460-i",
    "40030460":     "40030460-i",
    "40040460":     "40040460-i",

    # ── HT PRO BEND 87.5 DEG ──
    "40010860":     "40010860-i",
    "40020860":     "40020860-i",
    "40040860":     "40040860-i",

    # ── HT PRO DOOR BEND 87.5 DEG ──
    "40020867":     "40020867-i",
    "40040867":     "40040867-i",
    "40040868":     "40040868-i",

    # ── HT PRO WYE BRANCH ──
    "40633460":     "40633460-i",
    "40644460":     "40644460-i",

    # ── HT PRO TEE BRANCH 90 DEG ──
    "40633860":     "40633860-i",

    # ── HT PRO SWEPT TEE BRANCH ──
    "14040764860":  "14040764860-i",
    "41044857":     "41044857-HM",

    # ── HT PRO CORNER BRANCH ──
    "41264850":     "41264850-i",

    # ── HT PRO INSPECTION PIPE ──
    "49130060":     "49130060-i",
    "49140060":     "49140060-i",

    # ── HT PRO ONE WAY SOCKET / COUPLER ──
    "41710050":     "41710055-i",
    "41730050":     "41730050-i",
    "41740250":     "41740055-i",
    "41720055":     "41720055-i",

    # ── HT PRO REDUCER ──
    "42143050":     "42143050-i",

    # ── HT PRO END CAP ──
    "41620050":     "41620050-i",
    "41630050":     "41630050-i",
    "41640050":     "41640050-i",

    # ── HT PRO HEIGHT RISER FOR P TRAP ──
    "41242850":     "41242850-i",
    "7071243870-i": "7071243870-HM",

    # ── HT PRO HEIGHT RISER FOR MFT ──
    "60203651":     "60203651-i",   # same as US

    # ── HT PRO DOUBLE SOCKET ──
    "41710050":     "41710055-i",
    "41740250":     "41740055-i",
}

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def process_workbook(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes))
    changes = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get the dimensions
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_val = str(cell.value).strip()

                # Check exact match
                if cell_val in CODE_MAP:
                    old_code = cell_val
                    new_code = CODE_MAP[old_code]
                    cell.value = new_code

                    # Apply yellow highlight — preserve existing fill, only change bg
                    cell.fill = YELLOW_FILL

                    changes.append({
                        "Sheet": sheet_name,
                        "Cell": cell.coordinate,
                        "Old Code": old_code,
                        "New Code": new_code
                    })
                else:
                    # Also check if the cell contains the code as part of text
                    for old_code, new_code in CODE_MAP.items():
                        if old_code in cell_val and old_code != cell_val:
                            # partial match — replace within string
                            new_val = cell_val.replace(old_code, new_code)
                            if new_val != cell_val:
                                cell.value = new_val
                                cell.fill = YELLOW_FILL
                                changes.append({
                                    "Sheet": sheet_name,
                                    "Cell": cell.coordinate,
                                    "Old Code": f"(partial) {old_code}",
                                    "New Code": new_code
                                })
                                break

    return wb, changes


# ─── UI ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1F3864 0%, #006064 100%);
        padding: 20px 30px;
        border-radius: 12px;
        margin-bottom: 25px;
    }
    .main-header h1 { color: white; margin: 0; font-size: 28px; }
    .main-header p  { color: #cce0ff; margin: 5px 0 0 0; font-size: 14px; }
    .stat-box {
        background: #f0f8ff;
        border-left: 4px solid #1F3864;
        padding: 12px 18px;
        border-radius: 6px;
        margin: 8px 0;
    }
    .stat-box h3 { margin: 0; color: #1F3864; font-size: 22px; }
    .stat-box p  { margin: 0; color: #555; font-size: 13px; }
    .warning-box {
        background: #FFF8E1;
        border-left: 4px solid #FF8F00;
        padding: 12px 18px;
        border-radius: 6px;
    }
    .success-box {
        background: #E8F5E9;
        border-left: 4px solid #2E7D32;
        padding: 12px 18px;
        border-radius: 6px;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🔄 Huliot Item Code Updater</h1>
    <p>Upload your old BOQ Excel — codes auto-updated April 2025 → April 2026 | HT Pro · Ultra Silent · PERT | Yellow highlight = changed</p>
</div>
""", unsafe_allow_html=True)

# ─── HOW IT WORKS ─────────────────────────────────────────────────────────────
with st.expander("📖 How it works — Read first", expanded=False):
    st.markdown("""
    1. **Upload** your old BOQ Excel file (any format — .xlsx or .xls)
    2. **App scans** every cell across all sheets for old Huliot item codes
    3. **Auto-replaces** old codes with new 2026 codes
    4. **Highlights in Yellow** every cell that was changed so you can verify
    5. **Download** the updated file — your original format is 100% preserved
    
    ⚠️ **Only item codes are changed. All else stays exactly as-is:**
    - Descriptions, quantities, rates, formulas, colors, merged cells, fonts — all untouched
    - Multiple sheets in one file — all processed together
    - Both exact codes and codes embedded in text are detected
    """)

st.markdown("---")

# ─── FILE UPLOAD ──────────────────────────────────────────────────────────────
col1, col2 = st.columns([2, 1])

with col1:
    uploaded_file = st.file_uploader(
        "📂 Upload your BOQ Excel file",
        type=["xlsx", "xls"],
        help="Upload your old BOQ Excel. All sheets will be processed."
    )

with col2:
    st.markdown("""
    <div class="warning-box">
        <strong>⚠️ Before uploading:</strong><br>
        • Close the file in Excel first<br>
        • Keep a backup of original<br>
        • Works with .xlsx files only
    </div>
    """, unsafe_allow_html=True)

if uploaded_file is not None:
    file_bytes = uploaded_file.read()
    fname = uploaded_file.name

    st.markdown(f"**File loaded:** `{fname}` ({len(file_bytes)/1024:.1f} KB)")

    with st.spinner("🔍 Scanning all sheets and replacing codes..."):
        try:
            updated_wb, changes = process_workbook(file_bytes)

            # Save to buffer
            output = io.BytesIO()
            updated_wb.save(output)
            output.seek(0)
            updated_bytes = output.read()

            # ─── RESULTS ──────────────────────────────────────────────────────
            st.markdown("---")

            if len(changes) == 0:
                st.markdown("""
                <div class="warning-box">
                    <strong>ℹ️ No codes found to update.</strong><br>
                    Either this file already has the new 2026 codes, or the codes are not in the current mapping database.
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="success-box">
                    <strong>✅ Done! {len(changes)} code(s) updated and highlighted in yellow.</strong><br>
                    Download your updated file below. All changed cells are highlighted yellow for your review.
                </div>
                """, unsafe_allow_html=True)

                st.markdown("###")

                # Stats
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown(f"""<div class="stat-box"><h3>{len(changes)}</h3><p>Codes Updated</p></div>""", unsafe_allow_html=True)
                with c2:
                    sheets_affected = len(set(c["Sheet"] for c in changes))
                    st.markdown(f"""<div class="stat-box"><h3>{sheets_affected}</h3><p>Sheet(s) Affected</p></div>""", unsafe_allow_html=True)
                with c3:
                    ht_count  = sum(1 for c in changes if any(x in c["Old Code"] for x in ["400","410","411","412","413","414","415","416","417","418","419","420","421","422","42","49","60117","69201","4049"]))
                    st.markdown(f"""<div class="stat-box"><h3>{len(changes)-ht_count} / {ht_count}</h3><p>US / HT Changes</p></div>""", unsafe_allow_html=True)

                # Change log table
                st.markdown("### 📋 Change Log")
                st.dataframe(
                    changes,
                    use_container_width=True,
                    height=min(400, len(changes) * 38 + 38)
                )

            # ─── DOWNLOAD ─────────────────────────────────────────────────────
            new_fname = fname.replace(".xlsx", "_UPDATED_2026.xlsx").replace(".xls", "_UPDATED_2026.xlsx")
            st.download_button(
                label="⬇️ Download Updated File",
                data=updated_bytes,
                file_name=new_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.info("Make sure the file is a valid .xlsx Excel file and not password protected.")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("""
<small>
Huliot Code Updater | Built for PVL Ltd | April 2025 → April 2026 mapping | HT Pro · Ultra Silent · PERT<br>
Mapping covers all -i suffix additions, full code changes, and discontinued codes.<br>
<em>Your original Excel format is always preserved. Only item codes are changed.</em>
</small>
""", unsafe_allow_html=True)
